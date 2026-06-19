from dataclasses import dataclass
from pathlib import Path
from typing import Dict, Iterable, List

from openpyxl import load_workbook

ALLOWED_COURT_CODES = {"CA4", "CA5", "CA6", "CA7", "CA8", "CA9", "CA11"}


@dataclass(frozen=True)
class CourtUrlRow:
    row_number: int
    case_number: str
    court_code: str
    add_doc_url: str


def _clean(value) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _header_map(sheet) -> Dict[str, int]:
    """Return normalized header name to zero-based column index."""
    mapping: Dict[str, int] = {}
    for index, cell in enumerate(next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))):
        header = _clean(cell).lower()
        if header:
            mapping[header] = index
    return mapping


def read_court_urls(xlsm_path: str | Path) -> Dict[str, List[CourtUrlRow]]:
    """
    Read an .xlsm workbook and group rows by available court code.

    Required data:
      - CaseNumber: case/docket number to log with the opened URL.
      - LocationID: court code such as CA4, CA5, CA6, etc. Usually column C.
      - addDocURL: PACER URL to open. Usually column K.

    The reader first looks for the headers above, then falls back to the current
    known positions for LocationID and addDocURL when needed.
    """
    path = Path(xlsm_path).expanduser().resolve()
    if not path.exists():
        raise FileNotFoundError(f"Workbook not found: {path}")
    if path.suffix.lower() != ".xlsm":
        raise ValueError(f"Expected an .xlsm file, got: {path.name}")

    workbook = load_workbook(path, read_only=True, data_only=True, keep_vba=True)
    sheet = workbook.active
    headers = _header_map(sheet)

    # Header lookup, with fallback to the workbook layout already identified:
    # LocationID in C and addDocURL in K. CaseNumber is normally found by header.
    case_index = headers.get("casenumber")
    location_index = headers.get("locationid", 2)
    url_index = headers.get("adddocurl", 10)

    if case_index is None:
        raise ValueError("Could not find required header: CaseNumber")

    found: Dict[str, List[CourtUrlRow]] = {code: [] for code in sorted(ALLOWED_COURT_CODES)}

    for row_number, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
        case_number = _clean(row[case_index] if len(row) > case_index else "")
        location_id = _clean(row[location_index] if len(row) > location_index else "").upper()
        add_doc_url = _clean(row[url_index] if len(row) > url_index else "")

        if location_id in ALLOWED_COURT_CODES and add_doc_url:
            found[location_id].append(
                CourtUrlRow(
                    row_number=row_number,
                    case_number=case_number,
                    court_code=location_id,
                    add_doc_url=add_doc_url,
                )
            )

    return {court_code: rows for court_code, rows in found.items() if rows}


def select_courts_interactively(available: Dict[str, List[CourtUrlRow]]) -> List[str]:
    """Prompt the user to select court codes from the available workbook values."""
    if not available:
        return []

    print("\nAvailable court codes found in workbook:")
    for index, court_code in enumerate(available.keys(), start=1):
        print(f"  {index}. {court_code} ({len(available[court_code])} URL(s))")

    print("\nType court codes separated by commas, numbers separated by commas, or ALL.")
    print("Example: CA4,CA9 or 1,3")

    code_by_index = {str(index): code for index, code in enumerate(available.keys(), start=1)}

    while True:
        raw_selection = input("Select courts to open: ").strip()
        if not raw_selection:
            print("Please select at least one court code.")
            continue

        if raw_selection.upper() == "ALL":
            return list(available.keys())

        selected: List[str] = []
        invalid: List[str] = []
        for part in raw_selection.split(","):
            token = part.strip().upper()
            if not token:
                continue
            if token in code_by_index:
                selected.append(code_by_index[token])
            elif token in available:
                selected.append(token)
            else:
                invalid.append(token)

        if invalid:
            print(f"Invalid selection(s): {', '.join(invalid)}")
            continue

        if selected:
            return list(dict.fromkeys(selected))

        print("Please select at least one court code.")


def iter_selected_rows(
    grouped_rows: Dict[str, List[CourtUrlRow]],
    selected_courts: Iterable[str],
) -> Iterable[CourtUrlRow]:
    for court_code in selected_courts:
        yield from grouped_rows.get(court_code, [])
